"""Generate test plan Excel file."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Color palette ─────────────────────────────────────────────────────────────
C_HEADER_BG  = "1A3528"  # dark green
C_HEADER_FG  = "FFFFFF"
C_SECTION_BG = "243D2F"  # card green
C_SECTION_FG = "A8E040"  # lime
C_ROW_ALT    = "EAF4E3"  # light green tint
C_PASS       = "C6EFCE"
C_WARN       = "FFEB9C"
C_BORDER     = "2B4A38"

thin = Side(style="thin", color=C_BORDER)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_font(bold=True): return Font(name="Calibri", bold=bold, color=C_HEADER_FG, size=11)
def cell_font(bold=False): return Font(name="Calibri", bold=bold, size=10)
def section_font(): return Font(name="Calibri", bold=True, color=C_SECTION_FG, size=10)

def hdr_fill(): return PatternFill("solid", fgColor=C_HEADER_BG)
def section_fill(): return PatternFill("solid", fgColor=C_SECTION_BG)
def alt_fill(): return PatternFill("solid", fgColor=C_ROW_ALT)
def pass_fill(): return PatternFill("solid", fgColor=C_PASS)
def warn_fill(): return PatternFill("solid", fgColor=C_WARN)

def wrap(): return Alignment(wrap_text=True, vertical="top")
def center(): return Alignment(horizontal="center", vertical="center")

def write_header(ws, cols):
    for c, (title, width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=title)
        cell.font = hdr_font()
        cell.fill = hdr_fill()
        cell.alignment = center()
        cell.border = border
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[1].height = 22

def write_section(ws, row, label, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = section_font()
    cell.fill = section_fill()
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = border
    ws.row_dimensions[row].height = 18

def write_row(ws, row, values, alt=False, result_col=None, result_val=None):
    fill = alt_fill() if alt else None
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = cell_font(bold=(c == 1))
        cell.alignment = wrap()
        cell.border = border
        if result_col and c == result_col:
            if result_val and "pass" in str(result_val).lower():
                cell.fill = pass_fill()
            elif result_val and ("warn" in str(result_val).lower() or "lambat" in str(result_val).lower()):
                cell.fill = warn_fill()
            elif fill:
                cell.fill = fill
        elif fill:
            cell.fill = fill
    ws.row_dimensions[row].height = 30

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 1 — White Box Testing
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "White Box Testing"
ws1.freeze_panes = "A2"

cols = [("No", 12), ("Deskripsi", 52), ("Expected Results", 52), ("Status", 14)]
write_header(ws1, cols)

sections = [
    ("🔐 Autentikasi", [
        ("WB-AUTH-01", "Registrasi dengan email baru", "Akun dibuat, JWT token dikembalikan (HTTP 201)"),
        ("WB-AUTH-02", "Registrasi dengan email duplikat", 'Error "Email sudah terdaftar" (HTTP 400)'),
        ("WB-AUTH-03", "Registrasi dengan username duplikat", 'Error "Username sudah dipakai" (HTTP 400)'),
        ("WB-AUTH-04", "Registrasi dengan format email tidak valid", "Ditolak dengan pesan validasi (HTTP 422)"),
        ("WB-AUTH-05", "Login dengan kredensial valid", "Login berhasil, JWT token dikembalikan (HTTP 200)"),
        ("WB-AUTH-06", "Login dengan password salah", "Error 401"),
        ("WB-AUTH-07", "Login dengan email tidak terdaftar", "Error 401"),
        ("WB-AUTH-08", "Akses endpoint dengan token valid", "Request berhasil diproses (HTTP 200)"),
        ("WB-AUTH-09", "Akses endpoint tanpa token", "Ditolak dengan 401"),
        ("WB-AUTH-10", "Password tersimpan di database", "Tersimpan dalam bentuk hash bcrypt, bukan plaintext"),
        ("WB-AUTH-11", "Logout dari aplikasi", "Token dihapus dari storage dan state direset"),
        ("WB-AUTH-12", "Sesi berakhir (token expired)", "Redirect otomatis ke halaman login"),
    ]),
    ("👤 Profil & Target", [
        ("WB-PROF-01", "Lihat profil user", "Data profil tampil dengan benar"),
        ("WB-PROF-02", "Update profil (partial)", "Hanya field yang diisi yang berubah"),
        ("WB-PROF-03", "Hitung target nutrisi dengan profil lengkap", "Kalori dan makro terhitung dengan formula Mifflin-St Jeor"),
        ("WB-PROF-04", "Hitung target dengan profil belum lengkap", "Sistem memberi tahu profil perlu dilengkapi"),
        ("WB-PROF-05", 'Target untuk goal "lose"', "Kalori = TDEE dikurangi daily adjustment"),
        ("WB-PROF-06", 'Target untuk goal "gain"', "Kalori = TDEE ditambah daily adjustment"),
        ("WB-PROF-07", 'Target untuk goal "maintain"', "Kalori = TDEE"),
        ("WB-PROF-08", "Target dengan aktivitas custom", "TDEE = BMR + custom_exercise_calories"),
        ("WB-PROF-09", "Kalori hasil kalkulasi sangat rendah", "Ditahan di minimal 1200 kkal (floor)"),
        ("WB-PROF-10", "Forecast tersedia dengan profil lengkap", "Tanggal perkiraan target berat tercapai tampil"),
        ("WB-PROF-11", "Forecast dengan goal maintain", "Forecast tidak ditampilkan"),
        ("WB-PROF-12", "Update profil di app", "Target dan forecast otomatis diperbarui"),
    ]),
    ("🔍 Pencarian & Manajemen Makanan", [
        ("WB-FOOD-01", "Cari makanan yang ada di database lokal", "Hasil lokal muncul tanpa hit API eksternal"),
        ("WB-FOOD-02", "Cari makanan yang tidak ada di lokal", "USDA dipanggil dan hasil tampil"),
        ("WB-FOOD-03", "Custom food user muncul di hasil pencarian", "Custom food milik user tampil paling atas"),
        ("WB-FOOD-04", "Custom food user lain tidak muncul", "Hanya custom food milik user sendiri yang tampil"),
        ("WB-FOOD-05", "Buat custom food baru", "Custom food tersimpan dan muncul di daftar"),
        ("WB-FOOD-06", "Edit custom food", "Data custom food terupdate"),
        ("WB-FOOD-07", "API eksternal gagal/error", "Fallback ke hasil kosong, pencarian tetap jalan"),
        ("WB-FOOD-08", "Makanan dengan kalori 0 dari API", "Di-skip dan tidak masuk hasil pencarian"),
        ("WB-FOOD-09", "Cari dengan query kurang dari 2 karakter", "Pencarian tidak dijalankan"),
        ("WB-FOOD-10", "Log makanan dari USDA (tidak ada di DB lokal)", "Food item baru otomatis dibuat di DB (lazy insert)"),
    ]),
    ("📋 Log Makanan", [
        ("WB-LOG-01", "Log makanan berhasil", "Log tersimpan, kalori dan makro terhitung otomatis"),
        ("WB-LOG-02", "Lihat log harian", "Hanya log pada tanggal yang dipilih yang tampil"),
        ("WB-LOG-03", "Ringkasan nutrisi harian", "Total kalori, protein, karbs, lemak tampil (0 jika belum ada)"),
        ("WB-LOG-04", "Ringkasan menyertakan target", "Target harian ikut ditampilkan di ringkasan"),
        ("WB-LOG-05", "Edit jumlah gram log", "Kalori dan makro dihitung ulang otomatis"),
        ("WB-LOG-06", "Hapus log makanan", "Log terhapus dan summary terupdate"),
        ("WB-LOG-07", "Input gram 0 atau negatif", "Ditolak dengan pesan validasi"),
        ("WB-LOG-08", "Input gram lebih dari 10.000", "Ditolak dengan pesan validasi"),
        ("WB-LOG-09", "Log dengan unit selain gram (tbsp/tsp/cup)", "Dikonversi ke gram dengan benar sebelum disimpan"),
        ("WB-LOG-10", "Log dengan meal type berbeda", "Log masuk ke meal yang benar (breakfast/lunch/dinner/snack)"),
    ]),
    ("🔥 Streak & Riwayat", [
        ("WB-STR-01", "Hitung streak logging berturut-turut", "Jumlah hari berturut-turut terhitung dari hari ini ke belakang"),
        ("WB-STR-02", "Tidak ada log hari ini", "Streak = 0"),
        ("WB-STR-03", "Riwayat kalori 7 hari", "7 data tampil, hari tanpa log diisi 0"),
        ("WB-STR-04", "Riwayat menyertakan target kalori", "Target kalori per hari ikut tampil"),
    ]),
    ("⚖️ Log Berat Badan", [
        ("WB-WEIGHT-01", "Log berat badan hari ini", "Berat tersimpan dan profil user ikut terupdate"),
        ("WB-WEIGHT-02", "Log berat badan untuk tanggal lain", "Berat tersimpan, profil user tidak berubah"),
        ("WB-WEIGHT-03", "Log berat yang sudah ada (upsert)", "Data lama di-replace dan tidak duplikat"),
        ("WB-WEIGHT-04", "Lihat riwayat berat badan", "Riwayat tampil urut dari terlama, maksimal 30 entri"),
        ("WB-WEIGHT-05", "Log berat baru", "Forecast otomatis diperbarui"),
    ]),
    ("💧 Air Minum", [
        ("WB-WATER-01", "Lihat data air hari ini (belum ada log)", "Tampil 0 ml"),
        ("WB-WATER-02", "Update jumlah air minum", "Data tersimpan dan tampil terupdate"),
        ("WB-WATER-03", "Update air untuk tanggal yang sudah ada", "Data lama di-replace"),
        ("WB-WATER-04", "Target air untuk user pria", "Target = 3700 ml"),
        ("WB-WATER-05", "Target air untuk user wanita", "Target = 2700 ml"),
        ("WB-WATER-06", "User set target air custom", "Target custom digunakan, bukan default gender"),
    ]),
    ("💡 Insight Harian", [
        ("WB-INS-01", "Insight sudah ada hari ini", "Insight yang sama dikembalikan dari cache DB"),
        ("WB-INS-02", "Insight belum ada hari ini", "Insight baru di-generate dan disimpan ke DB"),
        ("WB-INS-03", "3+ hari berturut-turut kalori berlebih", "Insight mengingatkan soal kelebihan kalori konsisten"),
        ("WB-INS-04", "3+ hari berturut-turut kalori kurang", "Insight menyarankan tambah asupan"),
        ("WB-INS-05", "Hari tertentu konsisten surplus", "Insight menyebutkan pola hari dalam seminggu"),
        ("WB-INS-06", "Rata-rata 7 hari jauh dari target", "Insight menampilkan tren keseluruhan"),
        ("WB-INS-07", "Data log kurang dari 3 hari", "Insight mendorong user untuk terus logging"),
        ("WB-INS-08", "Tidak ada pola terdeteksi", 'Insight fallback "pola makanmu cukup stabil"'),
    ]),
    ("📅 Navigasi Diary (Flutter)", [
        ("WB-DIARY-01", "Navigasi ke hari sebelumnya", "Diary pindah ke kemarin dan data dimuat ulang"),
        ("WB-DIARY-02", "Navigasi ke hari berikutnya", "Diary pindah ke besok jika bukan hari ini"),
        ("WB-DIARY-03", "Navigasi maju dari hari ini", "Tombol next tidak bisa diklik"),
        ("WB-DIARY-04", "Label tanggal hari ini", 'Tampil "Today, MMM D"'),
        ("WB-DIARY-05", "Label tanggal kemarin", 'Tampil "Yesterday"'),
        ("WB-DIARY-06", "Label tanggal lain", 'Tampil format "D MMM YYYY"'),
    ]),
    ("🏠 Dashboard & UI (Flutter)", [
        ("WB-DASH-01", "Buka dashboard", "Semua data dimuat paralel (insight, target, streak, dll)"),
        ("WB-DASH-02", "Energy History chart tampil", "Stacked bar chart protein/karbs/lemak per hari tampil"),
        ("WB-DASH-03", "Bar chart tidak meluap container", "Tinggi bar dibatasi dan tidak overflow"),
        ("WB-DASH-04", "Tap bar di chart", "Tooltip muncul dengan detail tanggal dan breakdown makro"),
        ("WB-DASH-05", "Tunggu 3 detik setelah tap", "Tooltip hilang otomatis"),
        ("WB-DASH-06", "Status kalori — actual = 0", 'Tampil "START LOGGING TODAY"'),
        ("WB-DASH-07", "Status kalori — actual > 115% target", 'Tampil "TOO MUCH FOR TODAY"'),
        ("WB-DASH-08", "Status kalori — actual < 85% target", 'Tampil "BELOW TARGET TODAY"'),
        ("WB-DASH-09", "Status kalori — actual dalam range 85–115%", 'Tampil "ON TRACK TODAY"'),
        ("WB-DASH-10", "Forecast date label", 'Tampil dalam format "MMM DD, YYYY"'),
        ("WB-DASH-11", "Diary loading gagal atau timeout", "Loading indicator berhenti, app tidak stuck"),
    ]),
]

row = 2
alt = False
for section_label, items in sections:
    write_section(ws1, row, section_label, 4)
    row += 1
    for no, desc, expected in items:
        write_row(ws1, row, [no, desc, expected, ""], alt=alt)
        alt = not alt
        row += 1

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 2 — Non-Functional Testing
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Non-Functional Testing")
ws2.freeze_panes = "A2"

cols2 = [("No", 10), ("Endpoint", 36), ("Run 1", 12), ("Run 2", 12), ("Run 3", 12), ("Rata-rata", 14), ("Keterangan", 42)]
write_header(ws2, cols2)

perf_data = [
    ("NF-01", "POST /api/v1/auth/login",      "423ms", "392ms", "474ms", "430ms",  "Login + bcrypt verify"),
    ("NF-02", "GET /api/v1/foods/search",      "1356ms","1024ms","1145ms","1175ms", "Lambat karena hit USDA API eksternal"),
    ("NF-03", "GET /api/v1/logs/summary",      "180ms", "110ms", "102ms", "131ms",  "Agregasi DB harian"),
    ("NF-04", "GET /api/v1/logs/",             "190ms", "101ms", "107ms", "133ms",  "Daftar log harian"),
    ("NF-05", "GET /api/v1/insights/daily",    "197ms", "105ms", "101ms", "134ms",  "Cache hit (insight sudah ada)"),
    ("NF-06", "GET /api/v1/weight-logs/",      "107ms", "192ms", "106ms", "135ms",  "Riwayat berat 30 entri"),
    ("NF-07", "GET /api/v1/water/",            "103ms", "186ms", "108ms", "132ms",  "Data air harian"),
]

for i, row_data in enumerate(perf_data, 2):
    is_warn = "Lambat" in row_data[6]
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=i, column=c, value=val)
        cell.font = cell_font(bold=(c == 1))
        cell.alignment = wrap()
        cell.border = border
        if is_warn:
            cell.fill = warn_fill()
        elif i % 2 == 0:
            cell.fill = alt_fill()
    ws2.row_dimensions[i].height = 28

# Note row
note_row = len(perf_data) + 3
ws2.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=7)
note_cell = ws2.cell(row=note_row, column=1,
    value="Catatan: Semua endpoint diuji 3× dari client eksternal ke Railway deployment. Endpoint internal rata-rata <200ms (kategori Good, standar <500ms). "
          "Food search lambat karena memanggil USDA FoodData Central API secara real-time — bukan bottleneck internal.")
note_cell.font = Font(name="Calibri", italic=True, size=9, color="555555")
note_cell.alignment = Alignment(wrap_text=True, vertical="top")
ws2.row_dimensions[note_row].height = 42

# ─────────────────────────────────────────────────────────────────────────────
output = "/Users/toraaxlv/Desktop/Projects/Nutrishare/test_plan_tables.xlsx"
wb.save(output)
print(f"Saved: {output}")
